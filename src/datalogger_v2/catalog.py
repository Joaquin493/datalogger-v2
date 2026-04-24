"""Carga el catálogo de variables (entradas/salidas) desde el xlsx del PLC.

El xlsx es la "fuente de verdad" del mapeo: cada fila de la hoja (`Sheet2` por
defecto) representa una variable con dirección Schneider (`%I0.3` o `%Q2.5`),
símbolo legible (ej. `P_EMERG`) y descripción en castellano. Este módulo:

  - Parsea el archivo con openpyxl en modo read_only (streaming, poca RAM).
  - Filtra las filas a TYPE ∈ {INPUT, OUTPUT}; ignora REGISTER y TIMER.
  - Normaliza encoding CP1252 → UTF-8 en los comentarios (Ñ, acentos).
  - Asigna a cada variable un offset Modbus usando `addressing.assign_offsets`.

Devuelve una lista de `Variable` que el poller y la web consumen como catálogo.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from .addressing import ParsedAddress, assign_offsets, parse_address

log = logging.getLogger("datalogger_v2.catalog")


@dataclass(frozen=True)
class Variable:
    """Una variable del PLC ya resuelta y lista para registrar/mostrar."""
    address: str          # "%I0.3"
    type: str             # "INPUT" | "OUTPUT"
    symbol: str           # "P_EMERG"
    description: str      # "PARADA DE EMERGENCIA"
    kind: str             # "I" | "Q"
    offset: int           # offset Modbus dentro de su tabla (I=discrete inputs, Q=coils)


# Los comentarios vienen con caracteres CP1252 mal interpretados; reemplazo heurístico.
_ENCODING_FIXES = {
    "\ufffd": "Ñ",   # caso típico: SE�AL → SEÑAL
}


def _normalize(text: str | None) -> str:
    """Aplica los fixes de encoding y recorta espacios; None → ''."""
    if text is None:
        return ""
    out = str(text)
    for bad, good in _ENCODING_FIXES.items():
        out = out.replace(bad, good)
    return out.strip()


def load_catalog(
    xlsx_path: Path,
    sheet: str = "Sheet2",
    addressing_scheme: str = "sequential",
) -> list[Variable]:
    """Lee la hoja indicada y devuelve las variables TYPE in {INPUT, OUTPUT}.

    Formato esperado de columnas: POS | TYPE | Used | Address | Symbol | Comment.
    Se ignoran REGISTER y TIMER. No se filtra por columna 'Used'.
    Los offsets Modbus se calculan con `addressing.assign_offsets`.
    """
    log.info("Abriendo catálogo xlsx=%s hoja=%s scheme=%s", xlsx_path, sheet, addressing_scheme)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Hoja {sheet!r} no encontrada en {xlsx_path}")
    ws = wb[sheet]

    rows_raw: list[tuple[ParsedAddress, str, str, str]] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        _pos, typ, _used, addr, symbol, comment = (row + (None,) * 6)[:6]
        if typ not in ("INPUT", "OUTPUT"):
            continue
        if not addr:
            continue
        address = str(addr).strip()
        if address in seen:
            raise ValueError(f"Dirección duplicada en catálogo: {address}")
        seen.add(address)
        # Si no hay símbolo/descripción (I/O reservada), usamos la dirección como fallback.
        sym = str(symbol).strip() if symbol else address
        rows_raw.append((parse_address(address), typ, sym, _normalize(comment)))
    wb.close()

    offsets = assign_offsets([p for p, *_ in rows_raw], scheme=addressing_scheme)
    result = [
        Variable(
            address=p.raw,
            type=typ,
            symbol=sym,
            description=desc,
            kind=p.kind,
            offset=offsets[p.raw],
        )
        for (p, typ, sym, desc) in rows_raw
    ]
    n_i = sum(1 for v in result if v.kind == "I")
    n_q = sum(1 for v in result if v.kind == "Q")
    log.info("Catálogo cargado: %d variables (%d %%I, %d %%Q)", len(result), n_i, n_q)
    return result
